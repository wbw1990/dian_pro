
import os
import pdfplumber
import re
import math
import openpyxl
import json
import io
from config import * 
import numpy as np

import random

def get_filePath_fileName_fileExt(fileUrl):
    """
    获取文件路径， 文件名， 后缀名
    :param fileUrl:
    :return:
    """
    filepath, tmpfilename = os.path.split(fileUrl)
    shotname, extension = os.path.splitext(tmpfilename)
    return filepath, shotname, extension


def read_pdf(path):

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            txt = page.extract_text()
            # print(txt)
            # ss = page.bbox
            # tables = page.find_tables()


            name = re.compile('编号(.*?)\n',re.S).findall(txt)[0]
            name = name.replace(' ','')

            content = re.compile('4.工作任务\n(.*?)\n5.计划工作时间',re.S).findall(txt)[0]
            
            print(content)

            # for area_name in AREA_LIST:
            #     arr = content.split(area_name)
            #     if len(arr)>1:
               
            #         dict[area_name]


            result = []
            arr = content.split('\n')
            for a in arr:
                if '工作地点及设备双重名称' in a:
                    arr.remove(a)
                    continue

                device = ''
                # desc = ''

                a_arr = a.split('：')
                if len(a_arr)>1:
                    area = a_arr[0]
                    a_n_arr = a_arr[1].split(' ')
                    device =a_n_arr[0]
                    # if len(a_n_arr)>1:
                    #     desc = a_n_arr[1]

                    result.append({
                        'area': area,
                        'device':device
                    })
                else:
                    device_last = a_arr[0]
                    item = result.pop()
                    item['device'] = item['device'] + device_last
                    result.append(item)





            
            # 每页打印一分页分隔
            print('---------- ---------- ----------')
            break

        elec_contents_str = ''
        elec_contents = []
        for page in pdf.pages:
            txt = page.extract_text()            
            elec_content = re.compile('工作地点保留带电部分或注意事项.*?由工作票签发人填写(.*?)工作票签发人签名',re.S).findall(txt)
            
            if len(elec_content)==0:
                continue
            elec_contents_str = elec_content[0].replace(') （由工作许可人填写)','')
            elec_contents_str = elec_contents_str.replace('\n','')
           
            arr =elec_contents_str.split('； ')

            for a in arr:
                if '安全距离' in a:
                    continue
                if '工作中应做好相关协调配合工作' in a :
                    continue

                a_i =  re.split('[，、]', a)
                for a_ii in a_i:
                    elec_contents.append(a_ii)

            break
        
        print(result)
        return name,result,elec_content
       



# https://cloud.tencent.com/developer/ask/sof/78178
# https://www.jianshu.com/p/2381639ac6ef

def latlonhtoxyzwgs84(lat,lon,h):


    a=6378137.0             #radius a of earth in meters cfr WGS84
    b=6356752.3             #radius b of earth in meters cfr WGS84
    e2=1-(b**2/a**2)
    latr=lat/90*0.5*math.pi      #latitude in radians
    lonr=lon/180*math.pi         #longituede in radians
    Nphi=a/math.sqrt(1-e2*math.sin(latr)**2)
    x=(Nphi+h)*math.cos(latr)*math.cos(lonr)
    y=(Nphi+h)*math.cos(latr)*math.sin(lonr)
    z=(b**2/a**2*Nphi+h)*math.sin(latr)
    return x,y,z

def latlontoxyz(lat,lng,alt):
    phi = (90-lat)*(math.pi/180)
    theta = (lng+180)*(math.pi/180)
    radius = alt+200
    x = -(radius * math.sin(phi) * math.cos(theta))
    z = (radius * math.sin(phi) * math.sin(theta))
    y = (radius * math.cos(phi))
    x_ext= -713.349609375
    y_ext= 149.6861572265625
    z_ext= 166.3239288330078
    x = x + x_ext
    y = y + y_ext
    z = z + z_ext
    return x,y,z

def get_distince(x1,y1,z1,x2,y2,z2):
    # dis = np.sqrt(np.abs(x1-x2)* np.abs(x1-x2) + np.abs(y1-y2)* np.abs(y1-y2) + np.abs(z1-z2)* np.abs(z1-z2))
    dis = np.sqrt((x1-x2)* (x1-x2) + (y1-y2)* (y1-y2))
    return dis

# excel表格转json文件 https://www.jianshu.com/p/57a735ac021e
def excel_to_json(excel_file, json_file_name):
    # 加载工作薄
    book = openpyxl.load_workbook(excel_file)
    # 获取sheet页
    sheet = book["Sheet1"]
    # 行数
    max_row = sheet.max_row
    # 列数
    max_column = sheet.max_column
    print("max_row: %d, max_column: %d" % (max_row, max_column))
    # 结果，数组存储
    result = []
    heads = []
    max_column = 13
    # 解析表头
    for column in range(max_column):
        # 读取的话行列是从（1，1）开始
        heads.append(sheet.cell(1, column + 1).value)
    # 遍历每一行
    for row in range(max_row):
        if row == 0:
            continue
        one_line = {}
        for column in range(max_column):
            # 读取第二行开始每一个数据
            k = heads[column]
            cell = sheet.cell(row + 1, column + 1)
            value = cell.value
            if value == None:
                value =''
            if column ==3:
                value = value.replace('交流','')
            one_line[k] = value

        safe = get_safe(one_line['dianya'])
        if safe:
            one_line['person_safe'] = safe['person']
            one_line['car_safe'] = safe['car']
        else:
            one_line['person_safe'] = 0
            one_line['car_safe'] = 0

        print(one_line)
        result.append(one_line)
    book.close()
    # 将json保存为文件
    save_json_file(result, json_file_name)

# 将json保存为文件
def save_json_file(jd, json_file_name):
    file = io.open(json_file_name, 'w', encoding='utf-8')
    # 把对象转化为json对象
    # indent: 参数根据数据格式缩进显示，读起来更加清晰
    # ensure_ascii = True：默认输出ASCII码，如果把这个该成False, 就可以输出中文。
    txt = json.dumps(jd, indent=2, ensure_ascii=False)
    file.write(txt)
    file.close()




def generate_random_gps(base_log=None, base_lat=None, radius=None):
    radius_in_degrees = radius / 111300
    u = float(random.uniform(0.0, 1.0))
    v = float(random.uniform(0.0, 1.0))
    w = radius_in_degrees * math.sqrt(u)
    t = 2 * math.pi * v
    x = w * math.cos(t)
    y = w * math.sin(t)
    longitude = y + base_log
    latitude = x + base_lat
    # 这里是想保留6位小数点
    loga = '%.8f' % longitude
    lata = '%.8f' % latitude
    return float(loga), float(lata)


# log1, lat1 = generate_random_gps(base_log=120.7, base_lat=30, radius=1000000)

def load_json(path):
    with open(path,"r") as f:
        data = json.load(f)
        return data
    
if '__main__' == __name__:
     
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))   

    path = os.path.join(BASE_DIR,'piao','1000kV邢台站一次设备台账（2021.09.05）.xlsx')
    json_path = os.path.join(BASE_DIR,'pcd','tai.json')
     
    excel_to_json(path, json_path)


    # file_path = os.path.join(rundir,'piao','201909001.pdf') 
    # file_path2 = os.path.join(BASE_DIR,'piao','202104002.pdf') 
    # # result = read_pdf(file_path)
    # result2 = read_pdf(file_path2)
    # print(result2)


